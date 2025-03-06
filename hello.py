# Importing necessary libraries
import pandas as pd  # Import pandas for data manipulation (though it's not used in this code)
from openpyxl import load_workbook  # Import load_workbook from openpyxl to work with Excel files
import json  # Import json to handle JSON data
import xml.etree.ElementTree as ET  # Import ElementTree to work with XML data (though it's not used in this code)
from tabulate import tabulate  # Import tabulate to format data as a table for printing

# Load the Excel file into memory
wb = load_workbook("data_types.xlsx")  # Load the Excel file named 'data_types.xlsx'

# Extract Structured Data (Data stored in a structured format like tables)
sheet = wb["Structured Data"]  # Access the "Structured Data" sheet in the Excel workbook
data = list(sheet.iter_rows(values_only=True))  # Extract all rows of the sheet and convert them into a list of tuples
structured_data = [dict(zip(data[0], row)) for row in data[1:]]  # Create a list of dictionaries using the first row as keys and the subsequent rows as values
# The zip function pairs each key (from the first row) with each value (from the subsequent rows), 
# and dict converts this pairing into a dictionary.
# The list comprehension iterates through each row in the data (excluding the first row which is used as headers).

# Extract Semi-Structured Data (Data that might have different formats, like JSON and XML)
sheet = wb["Semi-Structured Data"]  # Access the "Semi-Structured Data" sheet in the Excel workbook
json_data, xml_data = sheet.cell(row=2, column=1).value, sheet.cell(row=4, column=1).value  # Get data from the 2nd and 4th rows, 1st column (JSON and XML data)
semi_structured_json = json.loads(json_data or '{}') if json_data else {}  # If there's JSON data, load it using json.loads(). If the data is empty (None), use an empty dictionary.
# The json.loads function parses the JSON string and converts it into a Python dictionary. 
# The or '{}' ensures that if json_data is None or empty, it defaults to an empty JSON string ('{}').

# Extract Unstructured Data (Data that is not easily categorized, like raw text or free-form data)
unstructured_data = wb["Unstructured Data"].cell(row=2, column=1).value  # Get the unstructured data from the 2nd row, 1st column of the "Unstructured Data" sheet

# Print categorized data
print("Structured data:\n", tabulate(structured_data, headers="keys", tablefmt="grid"))
# Print the structured data in a table format using the tabulate library.
# 'headers="keys"' uses the dictionary keys as headers, and 'tablefmt="grid"' formats the table as a grid.

print("\nSemi-structured data:\n", json.dumps(semi_structured_json, indent=4))
# Print the semi-structured JSON data, formatted with indentation for better readability.
# json.dumps() converts the Python dictionary back to a JSON string, and indent=4 adds indentation for readability.

print("\nUnstructured data:\n", unstructured_data)
# Print the unstructured data, which is assumed to be free-form text or another format not easily categorized.
